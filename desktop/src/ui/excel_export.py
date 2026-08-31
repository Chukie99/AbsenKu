"""
ui/excel_export.py — Unified Excel export hub.

Generate .xlsx for: siswa, absensi, nilai, ranking, jadwal, poin disiplin.
Wraps utils/excel_exporter.py with additional report generators.
"""
import os, sys
here = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(here))

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import db_manager as db
from utils import excel_exporter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

C = {
    "bg": "#F8F9FA",
    "card": "#FFFFFF",
    "accent": "#2563EB",
    "accent_light": "#EFF6FF",
    "green": "#16A34A",
    "green_light": "#F0FDF4",
    "red": "#DC2626",
    "red_light": "#FEF2F2",
    "orange": "#EA580C",
    "orange_light": "#FFF7ED",
    "purple": "#7C3AED",
    "purple_light": "#F5F3FF",
    "text": "#1E293B",
    "text2": "#64748B",
    "border": "#E2E8F0",
    "input_bg": "#F1F5F9",
}

# Style constants
HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F8F9FA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _styled_header(ws, headers, row=1):
    """Write a styled header row to a worksheet."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-size columns to content width."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


class Excel_export:
    def __init__(self, parent):
        self.parent = parent

    def build(self):
        p = self.parent
        for c in p.winfo_children():
            c.destroy()

        # ── Header ──
        hdr = tk.Frame(p, bg=C["accent"], height=70)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📊  Export Excel",
                font=("Segoe UI", 18, "bold"), fg="white",
                bg=C["accent"]).pack(anchor="w", padx=20, pady=14)

        # ── Content ──
        content = tk.Frame(p, bg=C["bg"])
        content.pack(fill="both", expand=True, padx=20, pady=20)

        tk.Label(content, text="Pilih data yang ingin di-export ke Excel:",
                font=("Segoe UI", 12), fg=C["text2"],
                bg=C["bg"]).pack(anchor="w", pady=(0, 16))

        # Cards grid
        grid = tk.Frame(content, bg=C["bg"])
        grid.pack(fill="both", expand=True)
        grid.columnconfigure(0, weight=1)
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(2, weight=1)

        cards = [
            (0, 0, "👤", "Data Siswa", "Export daftar lengkap siswa",
             C["accent"], self._export_siswa),
            (0, 1, "✅", "Absensi", "Export data absensi harian/bulanan",
             C["green"], self._export_absensi),
            (0, 2, "📝", "Nilai", "Export semua nilai siswa per mapel",
             C["orange"], self._export_nilai),
            (1, 0, "🏆", "Ranking", "Export ranking berdasarkan rata-rata",
             C["purple"], self._export_ranking),
            (1, 1, "📅", "Jadwal Pelajaran", "Export jadwal per kelas",
             C["red"], self._export_jadwal),
            (1, 2, "🎯", "Poin Disiplin", "Export poin disiplin siswa",
             C["orange"], self._export_poin),
        ]

        for row, col, icon, title, desc, color, cmd in cards:
            card = tk.Frame(grid, bg=C["card"], relief="flat",
                          highlightbackground=C["border"], highlightthickness=1)
            card.grid(row=row, column=col, sticky="nsew", padx=8, pady=8)
            grid.rowconfigure(row, weight=1)

            inner = tk.Frame(card, bg=C["card"])
            inner.pack(fill="both", expand=True, padx=16, pady=16)

            tk.Label(inner, text=icon, font=("Segoe UI", 28),
                    fg=color, bg=C["card"]).pack(anchor="w")
            tk.Label(inner, text=title, font=("Segoe UI", 13, "bold"),
                    fg=C["text"], bg=C["card"]).pack(anchor="w", pady=(4, 2))
            tk.Label(inner, text=desc, font=("Segoe UI", 10),
                    fg=C["text2"], bg=C["card"], wraplength=220).pack(
                        anchor="w", pady=(0, 10))
            tk.Button(inner, text="Export →", font=("Segoe UI", 10, "bold"),
                     bg=color, fg="white", relief="flat", cursor="hand2",
                     padx=16, pady=6, command=cmd).pack(anchor="w")

    # ── Export functions ──

    def _export_siswa(self):
        rows = db.siswa_all()
        if not rows:
            messagebox.showwarning("Info", "Tidak ada data siswa")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="data_siswa.xlsx")
        if not path:
            return
        try:
            kw = {k["id"]: k["nama"] for k in db.kelas_all()}
            enriched = []
            for r in rows:
                r["kelas_nama"] = kw.get(r.get("kelas_id"), "-")
                enriched.append(r)
            excel_exporter.export_siswa_xlsx(path, enriched)
            messagebox.showinfo("Sukses", f"Export {len(rows)} siswa ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export:\n{e}")

    def _export_absensi(self):
        top = tk.Toplevel(self.parent)
        top.title("Export Absensi")
        top.geometry("350x180")
        top.configure(bg=C["card"])
        top.resizable(False, False)
        top.grab_set()

        tk.Label(top, text="✅ Export Absensi",
                font=("Segoe UI", 13, "bold"),
                fg=C["text"], bg=C["card"]).pack(pady=(20, 12))

        frm = tk.Frame(top, bg=C["card"])
        frm.pack(fill="x", padx=20)

        tk.Label(frm, text="Tanggal (yyyy-mm-dd):", font=("Segoe UI", 10),
                fg=C["text2"], bg=C["card"]).pack(anchor="w")
        date_e = tk.Entry(frm, font=("Segoe UI", 11), bg=C["input_bg"],
                         relief="flat", highlightthickness=1,
                         highlightbackground=C["border"])
        date_e.pack(fill="x", ipady=4, pady=(2, 0))
        date_e.insert(0, datetime.now().strftime("%Y-%m-%d"))

        def generate():
            tgl = date_e.get().strip()
            rows = db.absensi_by_date(tgl)
            if not rows:
                messagebox.showwarning("Info",
                    f"Tidak ada data absensi {tgl}", parent=top)
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"absensi_{tgl}.xlsx",
                parent=top)
            if not path:
                return
            try:
                sw = {s["id"]: s for s in db.siswa_all()}
                mw = {m["id"]: m for m in db.mapel_all()}
                enriched = []
                for r in rows:
                    r["siswa_nama"] = sw.get(r.get("siswa_id"), {}).get("nama", "-")
                    r["mapel_nama"] = mw.get(r.get("mapel_id"), {}).get("nama", "-")
                    enriched.append(r)
                excel_exporter.export_absensi_xlsx(path, enriched)
                messagebox.showinfo("Sukses",
                    f"Export {len(rows)} data ke:\n{path}", parent=top)
                top.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Gagal export:\n{e}", parent=top)

        tk.Button(frm, text="Export Excel", font=("Segoe UI", 10, "bold"),
                 bg=C["green"], fg="white", relief="flat", cursor="hand2",
                 padx=16, pady=6, command=generate).pack(pady=(12, 0))

    def _export_nilai(self):
        rows = db.q("""
            SELECT s.nama, s.nis, m.nama AS mapel_nama,
                   n.nilai, n.semester, n.tahun_ajaran
            FROM nilai n
            JOIN siswa s ON n.siswa_id=s.id
            JOIN mapel m ON n.mapel_id=m.id
            ORDER BY s.nama, m.nama
        """)
        if not rows:
            messagebox.showwarning("Info", "Tidak ada data nilai")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="data_nilai.xlsx")
        if not path:
            return
        try:
            excel_exporter.export_nilai_xlsx(path, rows)
            messagebox.showinfo("Sukses", f"Export {len(rows)} data nilai ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export:\n{e}")

    def _export_ranking(self):
        rows = db.ranking_nilai_by_kelas()
        if not rows:
            messagebox.showwarning("Info", "Tidak ada data ranking")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="ranking_siswa.xlsx")
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ranking"
            headers = ["Rank", "NIS", "Nama", "Kelas", "Rata-rata Nilai", "Jumlah Mapel"]
            _styled_header(ws, headers)
            for i, r in enumerate(rows, 1):
                row_num = i + 1
                ws.cell(row_num, 1, i)
                ws.cell(row_num, 2, r.get("nis", ""))
                ws.cell(row_num, 3, r.get("nama", ""))
                ws.cell(row_num, 4, r.get("kelas_nama", ""))
                ws.cell(row_num, 5, round(r.get("rata_rata", 0), 1))
                ws.cell(row_num, 6, r.get("jumlah_mapel", 0))
                # Color code average
                avg = r.get("rata_rata", 0)
                if avg >= 80:
                    ws.cell(row_num, 5).font = Font(bold=True, color="16A34A")
                elif avg >= 60:
                    ws.cell(row_num, 5).font = Font(bold=True, color="EA580C")
                else:
                    ws.cell(row_num, 5).font = Font(bold=True, color="DC2626")
                # Alternate row shading
                if i % 2 == 0:
                    for c in range(1, 7):
                        ws.cell(row_num, c).fill = ALT_FILL
                for c in range(1, 7):
                    ws.cell(row_num, c).border = THIN_BORDER
            _auto_width(ws)
            wb.save(path)
            messagebox.showinfo("Sukses", f"Export {len(rows)} ranking ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export:\n{e}")

    def _export_jadwal(self):
        top = tk.Toplevel(self.parent)
        top.title("Export Jadwal")
        top.geometry("350x180")
        top.configure(bg=C["card"])
        top.resizable(False, False)
        top.grab_set()

        tk.Label(top, text="📅 Export Jadwal Pelajaran",
                font=("Segoe UI", 13, "bold"),
                fg=C["text"], bg=C["card"]).pack(pady=(20, 12))

        frm = tk.Frame(top, bg=C["card"])
        frm.pack(fill="x", padx=20)

        tk.Label(frm, text="Pilih Kelas:", font=("Segoe UI", 10),
                fg=C["text2"], bg=C["card"]).pack(anchor="w")
        kelas_list = db.kelas_all()
        kelas_map = {k["nama"]: k["id"] for k in kelas_list}
        kelas_cb = ttk.Combobox(frm, values=["Semua"] + [k["nama"] for k in kelas_list],
                               state="readonly", font=("Segoe UI", 10))
        kelas_cb.pack(fill="x", ipady=4, pady=(2, 0))
        kelas_cb.set("Semua")

        def generate():
            pilihan = kelas_cb.get()
            if pilihan == "Semua":
                jadwal = db.jadwal_all()
                filename = "jadwal_semua.xlsx"
            else:
                kid = kelas_map.get(pilihan)
                jadwal = db.jadwal_by_kelas(kid) if kid else []
                filename = f"jadwal_{pilihan}.xlsx"

            if not jadwal:
                messagebox.showwarning("Info",
                    "Tidak ada data jadwal", parent=top)
                return

            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=filename,
                parent=top)
            if not path:
                return
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Jadwal"
                headers = ["Hari", "Jam Mulai", "Jam Selesai", "Kelas",
                          "Mata Pelajaran", "Guru"]
                _styled_header(ws, headers)

                DAYS_ORDER = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu"]
                jadwal_sorted = sorted(jadwal,
                    key=lambda x: (
                        DAYS_ORDER.index(x.get("hari", "")) if x.get("hari", "") in DAYS_ORDER else 99,
                        x.get("jam_mulai", "")))

                for i, j in enumerate(jadwal_sorted, 2):
                    ws.cell(i, 1, j.get("hari", ""))
                    ws.cell(i, 2, j.get("jam_mulai", ""))
                    ws.cell(i, 3, j.get("jam_selesai", ""))
                    ws.cell(i, 4, j.get("kelas_nama", ""))
                    ws.cell(i, 5, j.get("mapel_nama", ""))
                    ws.cell(i, 6, j.get("guru") or "")
                    if i % 2 == 0:
                        for c in range(1, 7):
                            ws.cell(i, c).fill = ALT_FILL
                    for c in range(1, 7):
                        ws.cell(i, c).border = THIN_BORDER

                _auto_width(ws)
                wb.save(path)
                messagebox.showinfo("Sukses",
                    f"Export jadwal ke:\n{path}", parent=top)
                top.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Gagal export:\n{e}", parent=top)

        tk.Button(frm, text="Export Excel", font=("Segoe UI", 10, "bold"),
                 bg=C["purple"], fg="white", relief="flat", cursor="hand2",
                 padx=16, pady=6, command=generate).pack(pady=(12, 0))

    def _export_poin(self):
        rows = db.poin_summary_by_kelas()
        if not rows:
            messagebox.showwarning("Info", "Tidak ada data poin disiplin")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            initialfile="poin_disiplin.xlsx")
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Poin Disiplin"
            headers = ["Rank", "NIS", "Nama", "Kelas", "Poin Positif",
                      "Poin Negatif", "Poin Bersih"]
            _styled_header(ws, headers)
            for i, r in enumerate(rows, 1):
                row_num = i + 1
                ws.cell(row_num, 1, i)
                ws.cell(row_num, 2, r.get("nis", ""))
                ws.cell(row_num, 3, r.get("nama", ""))
                ws.cell(row_num, 4, r.get("kelas_nama", ""))
                ws.cell(row_num, 5, r.get("poin_positif", 0))
                ws.cell(row_num, 6, r.get("poin_negatif", 0))
                ws.cell(row_num, 7, r.get("poin_net", 0))
                # Color code net
                net = r.get("poin_net", 0)
                if net > 0:
                    ws.cell(row_num, 7).font = Font(bold=True, color="16A34A")
                elif net < 0:
                    ws.cell(row_num, 7).font = Font(bold=True, color="DC2626")
                if i % 2 == 0:
                    for c in range(1, 8):
                        ws.cell(row_num, c).fill = ALT_FILL
                for c in range(1, 8):
                    ws.cell(row_num, c).border = THIN_BORDER
            _auto_width(ws)
            wb.save(path)
            messagebox.showinfo("Sukses", f"Export {len(rows)} poin disiplin ke:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal export:\n{e}")
