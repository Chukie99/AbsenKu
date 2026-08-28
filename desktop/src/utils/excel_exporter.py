"""
excel_exporter.py — export tables to .xlsx via openpyxl.

Used by Report & Nilai screens for rapor-style Excel output.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def export_absensi_xlsx(path: str, rows: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Absensi"
    # header
    headers = ["Tanggal", "Waktu Masuk", "Waktu Keluar", "Status", "Siswa ID", "Mapel ID"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1A73E8")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    # rows
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row.get("tanggal"))
        ws.cell(r, 2, row.get("waktu_masuk"))
        ws.cell(r, 3, row.get("waktu_keluar"))
        ws.cell(r, 4, row.get("status"))
        ws.cell(r, 5, row.get("siswa_id"))
        ws.cell(r, 6, row.get("mapel_id"))
    # autosize
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    wb.save(path)
    return path


def export_nilai_xlsx(path: str, rows: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nilai"
    headers = ["Nama Siswa", "NIS", "Mapel", "Nilai", "Semester", "Tahun Ajaran"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A73E8")
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row.get("nama", ""))
        ws.cell(r, 2, row.get("nis", ""))
        ws.cell(r, 3, row.get("mapel_nama", ""))
        ws.cell(r, 4, row.get("nilai", ""))
        ws.cell(r, 5, row.get("semester", ""))
        ws.cell(r, 6, row.get("tahun_ajaran", ""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    wb.save(path)
    return path


def export_siswa_xlsx(path: str, rows: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Siswa"
    headers = ["NIS", "Nama", "Kelas", "Tgl Lahir", "Alamat", "No HP Ortu"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A73E8")
    for r, row in enumerate(rows, 2):
        ws.cell(r, 1, row.get("nis", ""))
        ws.cell(r, 2, row.get("nama", ""))
        ws.cell(r, 3, row.get("kelas_nama", ""))
        ws.cell(r, 4, row.get("tanggal_lahir", ""))
        ws.cell(r, 5, row.get("alamat", ""))
        ws.cell(r, 6, row.get("no_hp_ortu", ""))
    wb.save(path)
    return path
